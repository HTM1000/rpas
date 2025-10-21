# Verificar se há IDs duplicados no Excel
from openpyxl import load_workbook
from collections import Counter

EXCEL_FILE = "dados_processados.xlsx"

try:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Coluna ID
            ids.append(row[0])

    # Contar ocorrências
    counter = Counter(ids)
    duplicados = {id_val: count for id_val, count in counter.items() if count > 1}

    print("="*60)
    print("VERIFICAÇÃO DE DUPLICADOS NO EXCEL")
    print("="*60)
    print(f"\nTotal de registros no Excel: {len(ids)}")
    print(f"IDs únicos: {len(counter)}")

    if duplicados:
        print(f"\n[ERRO] DUPLICADOS ENCONTRADOS: {len(duplicados)}")
        print("\nIDs duplicados:")
        for id_val, count in duplicados.items():
            print(f"  - {id_val}: {count} vezes")
    else:
        print(f"\n[OK] NENHUM DUPLICADO! Todos os {len(ids)} registros tem IDs unicos!")

    print("="*60)

except FileNotFoundError:
    print(f"Arquivo {EXCEL_FILE} não encontrado!")
except Exception as e:
    print(f"Erro: {e}")
