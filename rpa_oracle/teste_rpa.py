# -*- coding: utf-8 -*-
"""
Script de Teste do RPA Oracle - MODO BACKGROUND
- Conecta na planilha de teste
- Busca linhas prontas para processar
- Gera arquivo Excel com os dados
- Gerencia cache de IDs processados
- Atualiza Status Oracle na planilha
- TESTE DE DUPLICAÇÃO: Executa múltiplas vezes para validar
"""
import json
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# Configurações
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
SPREADSHEET_ID = "147AN4Kn11T2qGyzTQgdqJ0QfSIt9TATEi0lw9zwMnpY"
SHEET_NAME = "Separação"

# Arquivos de controle
CACHE_FILE = "cache_teste.json"
EXCEL_FILE = "dados_processados.xlsx"
LOG_FILE = "teste_log.txt"

def log(msg):
    """Grava log em arquivo e imprime"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{timestamp}] {msg}"
    print(linha)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(linha + '\n')

def authenticate_google():
    """Autentica no Google Sheets"""
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("CredenciaisOracle.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())
    return build("sheets", "v4", credentials=creds)

def carregar_cache():
    """Carrega cache de IDs processados"""
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def salvar_cache(cache):
    """Salva cache de IDs processados"""
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, indent=2, ensure_ascii=False, fp=f)

def indice_para_coluna(idx):
    """Converte índice para letra de coluna (0=A, 1=B, etc)"""
    resultado = ""
    idx += 1  # Google Sheets é 1-based
    while idx > 0:
        idx -= 1
        resultado = chr(65 + (idx % 26)) + resultado
        idx //= 26
    return resultado

def buscar_linhas_para_processar(service):
    """Busca linhas com Status=CONCLUÍDO e Status Oracle vazio"""
    log("\n[BUSCA] Buscando linhas na planilha de teste...")

    res = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A1:AC"
    ).execute()

    valores = res.get("values", [])
    if not valores:
        log("[ERRO] Planilha vazia!")
        return [], None

    headers = valores[0]
    dados = valores[1:]

    log(f"[INFO] Colunas: {', '.join(headers)}")

    # Encontrar índices das colunas
    try:
        idx_id = headers.index("ID")
        idx_status = headers.index("Status")
        idx_status_oracle = headers.index("Status Oracle")
    except ValueError as e:
        log(f"[ERRO] Coluna não encontrada: {e}")
        return [], headers

    linhas = []
    for i, row in enumerate(dados):
        # Preencher linha curta
        if len(row) < len(headers):
            row += [''] * (len(headers) - len(row))

        id_item = row[idx_id].strip() if len(row) > idx_id else ""
        status = row[idx_status].strip().upper() if len(row) > idx_status else ""
        status_oracle = row[idx_status_oracle].strip() if len(row) > idx_status_oracle else ""

        # Critérios: Status contém CONCLUÍDO e Status Oracle vazio
        if "CONCLUÍDO" in status and status_oracle == "":
            linha_dict = dict(zip(headers, row))
            linhas.append((i + 2, linha_dict))  # +2 (header + 1-based)
            log(f"[OK] Linha {i+2}: ID={id_item}")

    log(f"[RESULTADO] {len(linhas)} linhas prontas para processar")
    return linhas, headers

def criar_excel():
    """Cria arquivo Excel se não existir"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Processados"
        ws.append(["ID", "Item", "Sub.Origem", "End. Origem", "Sub. Destino",
                   "End. Destino", "Quantidade", "Cód Referencia",
                   "Data Processamento", "Status", "Execução"])
        wb.save(EXCEL_FILE)
        log(f"[EXCEL] Arquivo criado: {EXCEL_FILE}")
    else:
        log(f"[EXCEL] Arquivo já existe: {EXCEL_FILE}")

def verificar_duplicacao_excel(id_item):
    """Verifica se ID já existe no Excel"""
    if not os.path.exists(EXCEL_FILE):
        return False

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == id_item:  # Coluna ID
            return True
    return False

def adicionar_ao_excel(linha, execucao):
    """Adiciona linha processada ao Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        linha.get("ID", ""),
        linha.get("Item", ""),
        linha.get("Sub.Origem", ""),
        linha.get("End. Origem", ""),
        linha.get("Sub. Destino", ""),
        linha.get("End. Destino", ""),
        linha.get("Quantidade", ""),
        linha.get("Cód Referencia", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Processado",
        execucao
    ])

    wb.save(EXCEL_FILE)
    log(f"[EXCEL] ID {linha.get('ID')} adicionado (Execução #{execucao})")

def atualizar_status_oracle(service, headers, linha_num, status_valor):
    """Atualiza Status Oracle na planilha"""
    try:
        idx_status_oracle = headers.index("Status Oracle")
        coluna_letra = indice_para_coluna(idx_status_oracle)
        range_str = f"{SHEET_NAME}!{coluna_letra}{linha_num}"

        log(f"[SHEETS] Atualizando {range_str} = '{status_valor}'")

        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=range_str,
            valueInputOption="RAW",
            body={"values": [[status_valor]]}
        ).execute()

        log(f"[SHEETS] Atualizado com sucesso!")
        return True
    except Exception as e:
        log(f"[ERRO] Falha ao atualizar Sheets: {e}")
        return False

def processar_linha(linha_dict, id_item, execucao):
    """Simula processamento de uma linha"""
    log(f"\n{'='*60}")
    log(f"[PROCESSANDO] ID: {id_item} (Execução #{execucao})")
    log(f"  Item: {linha_dict.get('Item', '')}")
    log(f"  Quantidade: {linha_dict.get('Quantidade', '')}")
    log(f"  Referência: {linha_dict.get('Cód Referencia', '')}")

    # Verificar duplicação no Excel
    if verificar_duplicacao_excel(id_item):
        log(f"[AVISO] ID {id_item} JÁ EXISTE NO EXCEL! (Tentativa de duplicação bloqueada)")
        return False

    # Simula delay de processamento
    time.sleep(0.5)

    # Adiciona ao Excel
    adicionar_ao_excel(linha_dict, execucao)
    log(f"[SUCESSO] Linha processada!")
    return True

def executar_ciclo(execucao):
    """Executa um ciclo completo de processamento"""
    log(f"\n{'#'*60}")
    log(f"# EXECUÇÃO #{execucao}")
    log(f"{'#'*60}")

    # Autenticar
    log("[1/5] Autenticando...")
    service = authenticate_google()

    # Criar Excel
    log("[2/5] Preparando Excel...")
    criar_excel()

    # Carregar cache
    log("[3/5] Carregando cache...")
    cache = carregar_cache()
    log(f"  Cache: {len(cache)} itens pendentes")

    # Buscar linhas
    log("[4/5] Buscando linhas...")
    linhas, headers = buscar_linhas_para_processar(service)

    if not linhas and not cache:
        log("[INFO] Nenhuma linha nova e nenhum pendente!")
        return 0, 0

    processadas = 0
    atualizadas = 0

    # Processar linhas novas
    if linhas:
        log(f"[5/5] Processando {len(linhas)} linhas...")
        for linha_num, linha_dict in linhas:
            id_item = linha_dict.get("ID", "").strip()

            # Verificar cache
            if id_item in cache:
                log(f"[CACHE] ID {id_item} já no cache. Pulando...")
                continue

            # Processar
            if processar_linha(linha_dict, id_item, execucao):
                processadas += 1

                # Adicionar ao cache como pendente
                cache[id_item] = {
                    "linha_atual": linha_num,
                    "execucao": execucao,
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                salvar_cache(cache)

                # Tentar atualizar Sheets
                if atualizar_status_oracle(service, headers, linha_num, "Processo Oracle Concluído"):
                    del cache[id_item]
                    salvar_cache(cache)
                    atualizadas += 1
                    log(f"[OK] ID {id_item} concluído!")
                else:
                    log(f"[PENDENTE] ID {id_item} ficou pendente")

    # Tentar atualizar pendentes do cache
    if cache:
        log(f"\n[RETRY] Tentando atualizar {len(cache)} pendentes...")
        for id_item, dados in list(cache.items()):
            linha_num = dados.get("linha_atual")
            if atualizar_status_oracle(service, headers, linha_num, "Processo Oracle Concluído"):
                del cache[id_item]
                salvar_cache(cache)
                atualizadas += 1
                log(f"[OK] ID {id_item} atualizado (retry)!")

    return processadas, atualizadas

def main():
    """Função principal - Executa múltiplos ciclos para testar duplicação"""
    log("="*60)
    log("TESTE DO RPA ORACLE - MODO BACKGROUND")
    log("="*60)

    # Limpar log anterior
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    total_processadas = 0
    total_atualizadas = 0

    # EXECUTAR 1 VEZ (já temos duplicados reais na planilha)
    NUM_EXECUCOES = 1

    for i in range(1, NUM_EXECUCOES + 1):
        processadas, atualizadas = executar_ciclo(i)
        total_processadas += processadas
        total_atualizadas += atualizadas

        if i < NUM_EXECUCOES:
            log(f"\n[AGUARDANDO] Pausa de 2s antes da próxima execução...")
            time.sleep(2)

    # Resumo final
    log("\n" + "="*60)
    log("RESUMO FINAL")
    log("="*60)
    log(f"Execuções realizadas: {NUM_EXECUCOES}")
    log(f"Total processadas: {total_processadas}")
    log(f"Total atualizadas no Sheets: {total_atualizadas}")

    # Verificar Excel
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        total_excel = ws.max_row - 1  # -1 para header
        log(f"Total de linhas no Excel: {total_excel}")

        if total_excel > total_processadas:
            log(f"[ERRO] DUPLICAÇÃO DETECTADA! Excel tem {total_excel - total_processadas} linhas a mais!")
        elif total_excel == total_processadas:
            log(f"[OK] Nenhuma duplicação! Excel correto.")

    # Verificar cache
    cache = carregar_cache()
    log(f"Itens ainda pendentes: {len(cache)}")

    log("="*60)
    log(f"\nArquivos gerados:")
    log(f"  - {EXCEL_FILE}")
    log(f"  - {CACHE_FILE}")
    log(f"  - {LOG_FILE}")
    log("\n[FIM] Teste concluído!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n[INTERROMPIDO] Teste cancelado")
    except Exception as e:
        log(f"\n[ERRO CRÍTICO] {e}")
        import traceback
        log(traceback.format_exc())
