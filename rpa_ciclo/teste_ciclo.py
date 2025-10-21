# -*- coding: utf-8 -*-
"""
Script de Teste do RPA_Ciclo - MODO TESTE COMPLETO
- Simula o fluxo completo: Oracle + Bancada
- Processa 50 itens por vez
- Loop contínuo: aguarda mais itens após processar
- Pula movimentos físicos do Oracle (apenas testa lógica)
- Integra com planilha de teste da bancada
"""
import json
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# =================== CONFIGURAÇÕES ===================
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Planilha de Separação (onde o RPA Oracle pega os dados)
SPREADSHEET_SEPARACAO_ID = "147AN4Kn11T2qGyzTQgdqJ0QfSIt9TATEi0lw9zwMnpY"
SHEET_SEPARACAO = "Separação"

# Planilha de Bancada (onde vamos buscar dados para teste)
SPREADSHEET_BANCADA_ID = "1KMS-1_FY6-cU26ZlaFu5jciSHEWlmluHo-QOFbB1LFE"
SHEET_BANCADA = "Bancada"

# Arquivos de controle
CACHE_FILE = "cache_teste_ciclo.json"
EXCEL_FILE_ORACLE = "dados_oracle_processados.xlsx"
EXCEL_FILE_BANCADA = "dados_bancada_processados.xlsx"
LOG_FILE = "teste_ciclo_log.txt"

# Limites
LIMITE_ITENS_POR_CICLO = 50
TEMPO_AGUARDO_SEM_ITENS = 30  # segundos

# =================== FUNÇÕES AUXILIARES ===================
def log(msg):
    """Grava log em arquivo e imprime"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{timestamp}] {msg}"
    print(linha)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(linha + '\n')

def authenticate_google():
    """Autentica no Google Sheets"""
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("CredenciaisOracle.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())
    return build("sheets", "v4", credentials=creds)

def carregar_cache():
    """Carrega cache de IDs processados"""
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def salvar_cache(cache):
    """Salva cache de IDs processados"""
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, indent=2, ensure_ascii=False, fp=f)

def indice_para_coluna(idx):
    """Converte índice para letra de coluna (0=A, 1=B, etc)"""
    resultado = ""
    idx += 1
    while idx > 0:
        idx -= 1
        resultado = chr(65 + (idx % 26)) + resultado
        idx //= 26
    return resultado

def criar_excel(arquivo, colunas):
    """Cria arquivo Excel se não existir"""
    if not os.path.exists(arquivo):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Processados"
        ws.append(colunas)
        wb.save(arquivo)
        log(f"[EXCEL] Arquivo criado: {arquivo}")
    else:
        log(f"[EXCEL] Arquivo já existe: {arquivo}")

def verificar_duplicacao_excel(arquivo, id_item):
    """Verifica se ID já existe no Excel"""
    if not os.path.exists(arquivo):
        return False

    wb = load_workbook(arquivo)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == id_item:
            return True
    return False

def adicionar_ao_excel(arquivo, linha_dados):
    """Adiciona linha processada ao Excel"""
    wb = load_workbook(arquivo)
    ws = wb.active
    ws.append(linha_dados)
    wb.save(arquivo)

# =================== ETAPA 1: RPA ORACLE (SIMULADO) ===================
def etapa_rpa_oracle(service, cache, ciclo):
    """
    Simula o RPA Oracle (sem movimentos físicos)
    - Busca até 50 linhas da planilha Separação
    - Valida regras
    - Adiciona ao Excel
    - Atualiza Status Oracle
    """
    log(f"\n{'='*60}")
    log(f"[RPA ORACLE] Iniciando processamento (Ciclo #{ciclo})")
    log(f"{'='*60}")

    # Buscar linhas para processar
    res = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_SEPARACAO_ID,
        range=f"{SHEET_SEPARACAO}!A1:AC"
    ).execute()

    valores = res.get("values", [])
    if not valores:
        log("[ORACLE] Planilha vazia!")
        return 0

    headers = valores[0]
    dados = valores[1:]

    # Encontrar índices das colunas
    try:
        idx_id = headers.index("ID")
        idx_status = headers.index("Status")
        idx_status_oracle = headers.index("Status Oracle")
    except ValueError as e:
        log(f"[ERRO] Coluna não encontrada: {e}")
        return 0

    # Filtrar linhas (Status=CONCLUÍDO e Status Oracle vazio)
    linhas_processar = []
    for i, row in enumerate(dados):
        if len(row) < len(headers):
            row += [''] * (len(headers) - len(row))

        id_item = row[idx_id].strip() if len(row) > idx_id else ""
        status = row[idx_status].strip().upper() if len(row) > idx_status else ""
        status_oracle = row[idx_status_oracle].strip() if len(row) > idx_status_oracle else ""

        if "CONCLUÍDO" in status and status_oracle == "":
            linha_dict = dict(zip(headers, row))
            linhas_processar.append((i + 2, linha_dict))

    log(f"[ORACLE] Encontradas {len(linhas_processar)} linhas para processar")

    if not linhas_processar:
        return 0

    # Limitar a 50 itens
    linhas_processar = linhas_processar[:LIMITE_ITENS_POR_CICLO]
    log(f"[ORACLE] Processando {len(linhas_processar)} linhas (limite: {LIMITE_ITENS_POR_CICLO})")

    # Criar Excel Oracle
    criar_excel(EXCEL_FILE_ORACLE, [
        "ID", "Item", "Sub.Origem", "End. Origem", "Sub. Destino",
        "End. Destino", "Quantidade", "Cód Referencia",
        "Data Processamento", "Status", "Ciclo"
    ])

    processadas = 0

    # Processar cada linha
    for linha_num, linha_dict in linhas_processar:
        id_item = linha_dict.get("ID", "").strip()
        item = linha_dict.get("Item", "")
        sub_o = linha_dict.get("Sub.Origem", "")
        end_o = linha_dict.get("End. Origem", "")
        sub_d = linha_dict.get("Sub. Destino", "")
        end_d = linha_dict.get("End. Destino", "")
        quantidade = linha_dict.get("Quantidade", "")
        referencia = linha_dict.get("Cód Referencia", "")

        # Verificar cache
        if id_item in cache:
            log(f"[ORACLE] ID {id_item} já processado (cache). Pulando...")
            continue

        # Verificar duplicação no Excel
        if verificar_duplicacao_excel(EXCEL_FILE_ORACLE, id_item):
            log(f"[ORACLE] ID {id_item} JÁ EXISTE NO EXCEL! Pulando...")
            continue

        # VALIDAÇÃO: ID não pode estar vazio
        if not id_item:
            log(f"[AVISO] Linha {linha_num} - ID vazio. Pulando.")
            continue

        # REGRA 1: Quantidade Zero
        try:
            qtd_float = float(str(quantidade).replace(",", ".").replace(" ", ""))
            if qtd_float == 0:
                atualizar_status_oracle(service, headers, linha_num, "Quantidade Zero")
                log(f"[REGRA 1] Linha {linha_num} - Quantidade Zero. Marcada.")
                continue
            elif qtd_float < 0:
                log(f"[AVISO] Linha {linha_num} - Quantidade negativa. Pulando.")
                continue
        except ValueError:
            log(f"[AVISO] Linha {linha_num} - Quantidade inválida. Pulando.")
            continue

        # REGRA 3: Campos vazios
        campos_obrigatorios = {
            "ITEM": item,
            "SUB. ORIGEM": sub_o,
            "END. ORIGEM": end_o,
            "SUB. DESTINO": sub_d,
            "END. DESTINO": end_d
        }

        # Para COD, não precisa verificar destino
        if str(referencia).strip().upper().startswith("COD"):
            campos_obrigatorios = {
                "ITEM": item,
                "SUB. ORIGEM": sub_o,
                "END. ORIGEM": end_o
            }

        campos_vazios = [nome for nome, valor in campos_obrigatorios.items() if not valor or str(valor).strip() == ""]
        if campos_vazios:
            atualizar_status_oracle(service, headers, linha_num, "Campo vazio encontrado")
            log(f"[REGRA 3] Linha {linha_num} - Campos vazios: {', '.join(campos_vazios)}")
            continue

        # REGRA 2 e 4: Transações não autorizadas
        subinvs_restritos = ["RAWINDIR", "RAWMANUT", "RAWWAFIFE"]
        sub_o_upper = str(sub_o).strip().upper()
        sub_d_upper = str(sub_d).strip().upper()

        if sub_o_upper in subinvs_restritos and sub_d_upper == "RAWCENTR":
            atualizar_status_oracle(service, headers, linha_num, "Transação não autorizada")
            log(f"[REGRA 2] Linha {linha_num} - Transação {sub_o} -> {sub_d} não autorizada")
            continue

        if sub_o_upper in subinvs_restritos and sub_o_upper == sub_d_upper:
            atualizar_status_oracle(service, headers, linha_num, "Transação não autorizada")
            log(f"[REGRA 4] Linha {linha_num} - Transação {sub_o} -> {sub_d} não autorizada")
            continue

        log(f"[ORACLE] Processando ID {id_item} | Item={item} | Qtd={quantidade}")

        # SIMULAÇÃO: Pular movimentos físicos do Oracle
        log(f"[MODO TESTE] Simulando preenchimento no Oracle (sem pyautogui)...")
        time.sleep(0.2)  # Simula tempo de processamento

        # Adicionar ao Excel
        adicionar_ao_excel(EXCEL_FILE_ORACLE, [
            id_item, item, sub_o, end_o, sub_d, end_d,
            quantidade, referencia,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Processado", ciclo
        ])

        # Adicionar ao cache
        cache[id_item] = {
            "linha_atual": linha_num,
            "ciclo": ciclo,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        salvar_cache(cache)

        # Atualizar Status Oracle no Sheets
        if atualizar_status_oracle(service, headers, linha_num, "Processo Oracle Concluído"):
            log(f"[ORACLE] ID {id_item} concluído!")
        else:
            log(f"[ORACLE] ID {id_item} processado, mas falhou ao atualizar Sheets")

        processadas += 1
        time.sleep(0.1)  # Pausa entre itens

    log(f"[ORACLE] Total processadas: {processadas}")
    return processadas

def atualizar_status_oracle(service, headers, linha, status_valor):
    """Atualiza o Status Oracle na planilha"""
    try:
        idx_status_oracle = headers.index("Status Oracle")
        coluna_letra = indice_para_coluna(idx_status_oracle)
        range_str = f"{SHEET_SEPARACAO}!{coluna_letra}{linha}"

        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_SEPARACAO_ID,
            range=range_str,
            valueInputOption="RAW",
            body={"values": [[status_valor]]}
        ).execute()

        return True
    except Exception as e:
        log(f"[ERRO] Falha ao atualizar Status Oracle: {e}")
        return False

# =================== ETAPA 2: RPA BANCADA (SIMULADO) ===================
def etapa_rpa_bancada(service, ciclo):
    """
    Simula o RPA Bancada
    - Busca dados da planilha de teste da bancada
    - Simula preenchimento na bancada
    - Adiciona ao Excel
    """
    log(f"\n{'='*60}")
    log(f"[RPA BANCADA] Iniciando processamento (Ciclo #{ciclo})")
    log(f"{'='*60}")

    try:
        # Buscar dados da planilha de bancada
        res = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_BANCADA_ID,
            range=f"{SHEET_BANCADA}!A1:Z"
        ).execute()

        valores = res.get("values", [])
        if not valores:
            log("[BANCADA] Planilha vazia!")
            return 0

        headers = valores[0]
        dados = valores[1:]

        log(f"[BANCADA] Colunas encontradas: {', '.join(headers)}")
        log(f"[BANCADA] Total de linhas: {len(dados)}")

        # Criar Excel Bancada
        criar_excel(EXCEL_FILE_BANCADA, headers + ["Data Processamento", "Ciclo"])

        processadas = 0

        # Processar cada linha
        for i, row in enumerate(dados[:LIMITE_ITENS_POR_CICLO], start=2):
            if len(row) < len(headers):
                row += [''] * (len(headers) - len(row))

            linha_dict = dict(zip(headers, row))

            # Simular preenchimento na bancada
            log(f"[BANCADA] Linha {i}: Simulando preenchimento...")
            time.sleep(0.1)  # Simula tempo de processamento

            # Adicionar ao Excel
            linha_dados = row + [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ciclo]
            adicionar_ao_excel(EXCEL_FILE_BANCADA, linha_dados)

            processadas += 1

        log(f"[BANCADA] Total processadas: {processadas}")
        return processadas

    except Exception as e:
        log(f"[ERRO BANCADA] {e}")
        import traceback
        log(traceback.format_exc())
        return 0

# =================== CICLO COMPLETO ===================
def executar_ciclo_completo(service, cache, ciclo):
    """Executa um ciclo completo: Oracle + Bancada"""
    log(f"\n{'#'*60}")
    log(f"# CICLO #{ciclo}")
    log(f"{'#'*60}")

    inicio = datetime.now()

    # Etapa 1: RPA Oracle
    itens_oracle = etapa_rpa_oracle(service, cache, ciclo)

    if itens_oracle == 0:
        log(f"[CICLO {ciclo}] Nenhum item processado no Oracle")
        return 0, 0

    # Pausa entre etapas
    log("[CICLO] Aguardando 2s entre Oracle e Bancada...")
    time.sleep(2)

    # Etapa 2: RPA Bancada
    itens_bancada = etapa_rpa_bancada(service, ciclo)

    fim = datetime.now()
    duracao = (fim - inicio).total_seconds()

    log(f"\n[RESUMO CICLO {ciclo}]")
    log(f"  Oracle: {itens_oracle} itens")
    log(f"  Bancada: {itens_bancada} itens")
    log(f"  Duração: {duracao:.2f}s")

    return itens_oracle, itens_bancada

# =================== MAIN LOOP ===================
def main():
    """Função principal - Loop contínuo"""
    log("="*60)
    log("TESTE DO RPA CICLO - MODO TESTE COMPLETO")
    log("="*60)
    log(f"Limite por ciclo: {LIMITE_ITENS_POR_CICLO} itens")
    log(f"Tempo de espera sem itens: {TEMPO_AGUARDO_SEM_ITENS}s")
    log("="*60)

    # Limpar log anterior
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    # Autenticar
    log("[1/2] Autenticando no Google Sheets...")
    service = authenticate_google()
    log("[1/2] Autenticação concluída!")

    # Carregar cache
    log("[2/2] Carregando cache...")
    cache = carregar_cache()
    log(f"[2/2] Cache: {len(cache)} itens processados anteriormente")

    ciclo = 0
    total_oracle = 0
    total_bancada = 0

    try:
        while True:
            ciclo += 1

            # Executar ciclo completo
            itens_oracle, itens_bancada = executar_ciclo_completo(service, cache, ciclo)

            total_oracle += itens_oracle
            total_bancada += itens_bancada

            # Se não processou nada, aguardar
            if itens_oracle == 0:
                log(f"\n[AGUARDANDO] Nenhum item novo. Aguardando {TEMPO_AGUARDO_SEM_ITENS}s...")
                log("[INFO] O loop continuará verificando automaticamente...")
                time.sleep(TEMPO_AGUARDO_SEM_ITENS)
                continue

            # Pausa breve entre ciclos
            log(f"\n[PRÓXIMO CICLO] Aguardando 5s antes do próximo ciclo...")
            time.sleep(5)

    except KeyboardInterrupt:
        log("\n[INTERROMPIDO] Teste cancelado pelo usuário (Ctrl+C)")
    except Exception as e:
        log(f"\n[ERRO CRÍTICO] {e}")
        import traceback
        log(traceback.format_exc())
    finally:
        # Resumo final
        log("\n" + "="*60)
        log("RESUMO FINAL")
        log("="*60)
        log(f"Total de ciclos executados: {ciclo}")
        log(f"Total Oracle: {total_oracle} itens")
        log(f"Total Bancada: {total_bancada} itens")
        log("="*60)
        log(f"\nArquivos gerados:")
        log(f"  - {EXCEL_FILE_ORACLE}")
        log(f"  - {EXCEL_FILE_BANCADA}")
        log(f"  - {CACHE_FILE}")
        log(f"  - {LOG_FILE}")
        log("\n[FIM] Teste concluído!")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"\n[ERRO FATAL] {e}")
        import traceback
        log(traceback.format_exc())
